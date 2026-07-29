from collections import defaultdict
from datetime import date, timedelta
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse

from sqlalchemy import func
from sqlalchemy.orm import Session
from sqlalchemy import extract
from datetime import date, timedelta
from app.models.daily_report import DailyReport
from app.models.store import Store
from openpyxl import Workbook
from openpyxl.styles import Font

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
)

from app.database import get_db

from app.dependencies.auth import get_current_user
from app.dependencies.roles import require_role

from app.models.daily_report import DailyReport
from app.models.expense import Expense
from app.models.purchase import Purchase
from app.models.store import Store
from app.models.udhaar_entry import UdhaarEntry


router = APIRouter(
    prefix="/analytics",
    tags=["Analytics"],
)


def _safe_number(value):

    if value is None:
        return 0

    return float(value)

from datetime import datetime

def _get_period_bounds(
    period,
    from_date=None,
    to_date=None,
):
    today = date.today()

    if period == "today":
        return today, today

    elif period in ["7days", "last7"]:
        return today - timedelta(days=7), today

    elif period in ["30days", "last30"]:
        return today - timedelta(days=30), today

    elif period == "90days":
        return today - timedelta(days=90), today

    elif period in ["month", "thisMonth"]:
        return today.replace(day=1), today

    elif period == "last_month":
        first_this_month = today.replace(day=1)
        last_day_prev = first_this_month - timedelta(days=1)
        first_prev = last_day_prev.replace(day=1)
        return first_prev, last_day_prev

    elif period in ["year", "thisYear"]:
        return date(today.year, 1, 1), today

    elif period == "custom":
        start = (
            datetime.strptime(from_date, "%Y-%m-%d").date()
            if from_date else None
        )

        end = (
            datetime.strptime(to_date, "%Y-%m-%d").date()
            if to_date else None
        )

        return start, end

    return None, None


def _get_reports(
    db,
    period=None,
    store_id=None,
    from_date=None,
    to_date=None,
    submitted_only=False,
) -> List[DailyReport]:

    query = db.query(DailyReport)

    # Only include submitted reports when required
    if submitted_only:
        query = query.filter(
            DailyReport.is_submitted == True
        )

    start_date, end_date = _get_period_bounds(
        period,
        from_date,
        to_date,
    )

    if start_date:
        query = query.filter(
            DailyReport.report_date >= start_date
        )

    if end_date:
        query = query.filter(
            DailyReport.report_date <= end_date
        )

    if (
        store_id
        and str(store_id).lower() != "all"
    ):
        query = query.filter(
            DailyReport.store_id == int(store_id)
        )

    return query.all()

def _monthly_series(
    reports: List[DailyReport],
):

    monthly = defaultdict(float)

    for report in reports:

        if not report.report_date:
            continue

        month = report.report_date.strftime("%b")

        revenue = (
            _safe_number(report.cash_sales)
            + _safe_number(report.upi_sales)
            + _safe_number(report.card_sales)
            + _safe_number(report.udhaar_sales)
        )

        monthly[month] += revenue

    return [
        {
            "month": month,
            "revenue": value,
        }
        for month, value in monthly.items()
    ]

def _get_outstanding_entries(
    db: Session,
    period=None,
    store_id="all",
):

    query = db.query(UdhaarEntry)

    if (
        store_id
        and str(store_id).lower() != "all"
    ):
        query = query.filter(
            UdhaarEntry.store_id == int(store_id)
        )

    return query.all()

def _today_report(db: Session, store_id: int):

    return (
        db.query(DailyReport)
        .filter(
            DailyReport.store_id == store_id,
            DailyReport.report_date == date.today(),
        )
        .first()
    )


@router.get("/dashboard-summary")
def dashboard_summary(
    period: str = "today",
    store_id: str = "all",
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):

    require_role(
        ["owner", "store_manager"],
        current_user["role"],
    )

    if current_user["role"] == "store_manager":
        store_id = str(current_user["store_id"])

    reports = _get_reports(
    db=db,
    period=period,
    store_id=store_id,
    submitted_only=current_user["role"] == "owner",
)
    total_sales = 0
    total_bills = 0
    total_deliveries = 0
    total_udhaar = 0
    total_purchases = 0

    for report in reports:

        total_sales += (
            _safe_number(report.cash_sales)
        + _safe_number(report.upi_sales)
        + _safe_number(report.card_sales)
        + _safe_number(report.udhaar_sales)
    )

        total_purchases += _safe_number(
        report.total_purchases
    )

        total_bills += report.total_bills or 0

        total_deliveries += report.deliveries or 0

        total_udhaar += _safe_number(
        report.udhaar_sales
    )

# ----------------------------------------------------
# Expenses are calculated from the Expense table
# ----------------------------------------------------

    expense_query = db.query(func.sum(Expense.amount))

    if store_id != "all":
        expense_query = expense_query.filter(
        Expense.store_id == int(store_id)
    )

    start_date, end_date = _get_period_bounds(period)

    if start_date:
        expense_query = expense_query.filter(
        func.date(
    func.datetime(
        Expense.created_at,
        "+5 hours",
        "+30 minutes",
    )
) >= start_date
    )

    if end_date:
        expense_query = expense_query.filter(
        func.date(
    func.datetime(
        Expense.created_at,
        "+5 hours",
        "+30 minutes",
    )
) <= end_date
    )

    total_expenses = expense_query.scalar() or 0

    

    average_bill = (
        total_sales / total_bills
        if total_bills
        else 0
    )

    return {

        "total_sales": round(total_sales, 2),

        "total_revenue": round(total_sales, 2),

        "total_purchases": round(total_purchases,2,),

        "total_deliveries": total_deliveries,

        "total_bills": total_bills,

        "average_bill_value": round(
            average_bill,
            2,
        ),

        "average_bill": round(
            average_bill,
            2,
        ),

        "total_expenses": round(
            total_expenses,
            2,
        ),

        "total_udhaar": round(
            total_udhaar,
            2,
        ),

        "outstanding_udhaar": round(
            total_udhaar,
            2,
        ),

        "growth_rate": 0,

        "submitted_reports": len(reports),
    }

@router.get("/store-summary")
def store_summary(
    period: str = "today",
    store_id: str = "all",
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):

    require_role(
        ["owner", "store_manager"],
        current_user["role"],
    )

    if current_user["role"] == "store_manager":
        store_id = str(current_user["store_id"])

    reports = _get_reports(
        db=db,
        period=period,
        store_id=store_id,
        submitted_only=current_user["role"] == "owner",
    )

    stores = {}

    for report in reports:

        sid = report.store_id

        if sid not in stores:

            store = (
                db.query(Store)
                .filter(Store.id == sid)
                .first()
            )

            stores[sid] = {
                "store_id": sid,
                "store_name": (
                    store.name
                    if store
                    else f"Store {sid}"
                ),
                "total_sales": 0,
                "total_bills": 0,
                "total_expenses": 0,
                "total_purchases": 0,
            }

        stores[sid]["total_sales"] += (
            _safe_number(report.cash_sales)
            + _safe_number(report.upi_sales)
            + _safe_number(report.card_sales)
            + _safe_number(report.udhaar_sales)
        )

        stores[sid]["total_bills"] += (
            report.total_bills or 0
        )

        stores[sid]["total_expenses"] += (
            _safe_number(report.total_expenses)
        )

        stores[sid]["total_purchases"] += (
            _safe_number(report.total_purchases)
        )

    for sid in stores:

        stores[sid]["total_sales"] = round(
            stores[sid]["total_sales"],
            2,
        )

        stores[sid]["total_expenses"] = round(
            stores[sid]["total_expenses"],
            2,
        )

        stores[sid]["total_purchases"] = round(
            stores[sid]["total_purchases"],
            2,
        )

        stores[sid]["growth_rate"] = 0

    return list(stores.values())

@router.get("/payment-breakdown")
def payment_breakdown(
    period: str = "today",
    store_id: str = "all",
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):

    require_role(
        ["owner", "store_manager"],
        current_user["role"],
    )

    if current_user["role"] == "store_manager":
        store_id = str(current_user["store_id"])

    reports = _get_reports(
    db=db,
    period=period,
    store_id=store_id,
    submitted_only=current_user["role"] == "owner",
)
    cash = 0
    upi = 0
    card = 0
    udhaar = 0

    for report in reports:

        cash += _safe_number(report.cash_sales)
        upi += _safe_number(report.upi_sales)
        card += _safe_number(report.card_sales)
        udhaar += _safe_number(report.udhaar_sales)

    return [
        {
            "name": "Cash",
            "value": round(cash, 2),
        },
        {
            "name": "UPI",
            "value": round(upi, 2),
        },
        {
            "name": "Card",
            "value": round(card, 2),
        },
        {
            "name": "Udhaar",
            "value": round(udhaar, 2),
        },
    ]

@router.get("/expense-distribution")
def expense_distribution(
    period: str = "today",
    store_id: str = "all",
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    require_role(
        ["owner", "store_manager"],
        current_user["role"],
    )

    if current_user["role"] == "store_manager":
        store_id = str(current_user["store_id"])

    query = db.query(Expense)

    # ---------- Store Filter ----------
    if store_id != "all":
        query = query.filter(
            Expense.store_id == int(store_id)
        )

    # ---------- Period Filter ----------
    start_date, end_date = _get_period_bounds(period)

    print("Start Date:", start_date)
    print("End Date:", end_date)
    print("Store ID:", store_id)

    if start_date:
        query = query.filter(
            func.date(
    func.datetime(
        Expense.created_at,
        "+5 hours",
        "+30 minutes",
    )
) >= start_date
        )

    if end_date:
        query = query.filter(
            func.date(
    func.datetime(
        Expense.created_at,
        "+5 hours",
        "+30 minutes",
    )
) <= end_date
        )

    # ---------- Debug: All Expenses ----------
    all_expenses = db.query(Expense).all()

    print("\n===== ALL EXPENSES =====")

    for expense in all_expenses:
        print(
            expense.id,
            expense.expense_type,
            expense.amount,
            expense.created_at,
        )

    # ---------- Filtered Expenses ----------
    expenses = query.all()

    print("\n===== FILTERED EXPENSES =====")
    print("Expenses Found:", len(expenses))

    for expense in expenses:
        print(
            expense.id,
            expense.expense_type,
            expense.amount,
            expense.created_at,
        )

    grouped = defaultdict(float)

    for expense in expenses:
        grouped[expense.expense_type] += _safe_number(
            expense.amount
        )

    return [
        {
            "name": name,
            "amount": round(amount, 2),
        }
        for name, amount in grouped.items()
    ]
@router.get("/sales-trend")
def sales_trend(
    period: str = "today",
    store_id: str = "all",
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):

    require_role(
        ["owner", "store_manager"],
        current_user["role"],
    )

    if current_user["role"] == "store_manager":
        store_id = str(current_user["store_id"])

    reports = _get_reports(
    db=db,
    period=period,
    store_id=store_id,
    submitted_only=current_user["role"] == "owner",
)

    return _monthly_series(reports)

@router.get("/outstanding-udhaar")
def outstanding_udhaar(
    period: str = "today",
    store_id: str = "all",
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    require_role(
        ["owner", "store_manager"],
        current_user["role"],
    )

    if current_user["role"] == "store_manager":
        store_id = str(current_user["store_id"])

    query = db.query(
        Store.name.label("store_name"),
        func.sum(DailyReport.udhaar_sales).label("outstanding"),
    ).join(
        DailyReport,
        DailyReport.store_id == Store.id,
    )

    # ---------- Store Filter ----------
    if store_id != "all":
        query = query.filter(
            DailyReport.store_id == int(store_id)
        )

    # ---------- Period Filter ----------
    today = date.today()

    if period == "today":
        query = query.filter(
            DailyReport.report_date == today
        )

    elif period == "last7":
        query = query.filter(
            DailyReport.report_date >= today - timedelta(days=7)
        )

    elif period == "last30":
        query = query.filter(
            DailyReport.report_date >= today - timedelta(days=30)
        )

    elif period == "thisMonth":
        query = query.filter(
            extract("month", DailyReport.report_date) == today.month,
            extract("year", DailyReport.report_date) == today.year,
        )

    elif period == "thisYear":
        query = query.filter(
            extract("year", DailyReport.report_date) == today.year,
        )

    rows = (
        query.group_by(Store.name)
        .all()
    )

    return [
        {
            "store_name": row.store_name,
            "pending": float(row.outstanding or 0),
            "recovered": 0,
            "outstanding": float(row.outstanding or 0),
            "recovery_rate": 0,
        }
        for row in rows
    ]




@router.get("/performance")
def performance(
    period: str = "today",
    store_id: str = "all",
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):

    require_role(
        ["owner", "store_manager"],
        current_user["role"],
    )

    if current_user["role"] == "store_manager":
        store_id = str(current_user["store_id"])

    reports = _get_reports(
    db=db,
    period=period,
    store_id=store_id,
    submitted_only=current_user["role"] == "owner",
)

    revenue = sum(

        _safe_number(report.cash_sales)

        + _safe_number(report.upi_sales)

        + _safe_number(report.card_sales)

        + _safe_number(report.udhaar_sales)

        for report in reports

    )

    expenses = sum(
        _safe_number(report.total_expenses)
        for report in reports
    )

    profit = revenue - expenses

    margin = (
        (profit / revenue) * 100
        if revenue
        else 0
    )

    return {

        "revenue": round(revenue, 2),

        "expenses": round(expenses, 2),

        "profit": round(profit, 2),

        "profit_margin": round(
            margin,
            1,
        ),

    }

@router.get("/manager-hero")
def manager_hero(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    require_role(
        ["store_manager"],
        current_user["role"],
    )

    today = date.today()

    report = (
        db.query(DailyReport)
        .filter(
            DailyReport.store_id == current_user["store_id"],
            DailyReport.report_date == today,
        )
        .first()
    )

    if report is None:
        report = DailyReport(
            store_id=current_user["store_id"],
            submitted_by=current_user["user_id"],
            report_date=today,
        )

        db.add(report)
        db.commit()
        db.refresh(report)

    purchases_count = (
        db.query(Purchase)
        .filter(
            Purchase.store_id == current_user["store_id"],
            func.date(Purchase.purchase_date) == report.report_date,
        )
        .count()
    )

    

    return {
        "user": {
            "full_name": current_user["full_name"],
            "role": current_user["role"],
            "store_id": current_user["store_id"],
        },
        "report": {
            "status": "Locked" if report.is_locked else "In Progress",

            "sales_completed": (
                (report.cash_sales or 0)
                + (report.upi_sales or 0)
                + (report.card_sales or 0)
                + (report.udhaar_sales or 0)
            ) > 0,

            "expenses_completed": (report.total_expenses or 0) > 0,

            "purchases_completed": (
                (report.total_purchases or 0) > 0
                or purchases_count > 0
            ),

            "deliveries_completed": (report.deliveries or 0) > 0,

            

            "notes_completed": bool(report.notes),
        },
    }

@router.get("/manager-dashboard")
def manager_dashboard(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    require_role(
        ["store_manager"],
        current_user["role"],
    )

    store_id = current_user["store_id"]
    today = date.today()

    report = (
        db.query(DailyReport)
        .filter(
            DailyReport.store_id == store_id,
            DailyReport.report_date == today,
        )
        .first()
    )

    purchases = (
        db.query(Purchase)
        .filter(
            Purchase.store_id == store_id,
            func.date(Purchase.purchase_date) == today,
        )
        .order_by(Purchase.purchase_date.desc())
        .all()
    )

    expenses = (
    db.query(Expense)
    .filter(
        Expense.store_id == store_id,
        func.date(
    func.datetime(
        Expense.created_at,
        "+5 hours",
        "+30 minutes",
    )
) == today,
    )
    .order_by(Expense.created_at.desc())
    .all()
)
    
    payment_breakdown = {
        "cash": report.cash_sales if report else 0,
        "upi": report.upi_sales if report else 0,
        "card": report.card_sales if report else 0,
        "udhaar": report.udhaar_sales if report else 0,
    }

    progress = {
        "sales_completed":
            (
                payment_breakdown["cash"]
                + payment_breakdown["upi"]
                + payment_breakdown["card"]
                + payment_breakdown["udhaar"]
            ) > 0,

        "expenses_completed": len(expenses) > 0,

        "purchases_completed": len(purchases) > 0,

        "deliveries_completed":
            report.deliveries > 0 if report else False,


        "report_submitted":
            report.is_locked if report else False,
    }

    return {

        "progress": progress,

        "payment_breakdown": payment_breakdown,

        "purchases": [
            {
                "product_name": p.product_name,
                "supplier_name": p.supplier_name,
                "quantity": p.quantity,
                "amount": p.purchase_amount,
            }
            for p in purchases
        ],

        "expenses": [
            {
                "expense_type": e.expense_type,
                "amount": e.amount,
            }
            for e in expenses
        ],

        
    }

@router.get("/overview")
def overview(
    period: str = "today",
    store_id: str = "all",
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):

    require_role(
        ["owner", "store_manager"],
        current_user["role"],
    )

    if current_user["role"] == "store_manager":
        store_id = str(current_user["store_id"])

    reports = _get_reports(
    db=db,
    period=period,
    store_id=store_id,
    submitted_only=current_user["role"] == "owner",
)

    sales = sum(

        _safe_number(report.cash_sales)

        + _safe_number(report.upi_sales)

        + _safe_number(report.card_sales)

        + _safe_number(report.udhaar_sales)

        for report in reports

    )

    expenses = sum(
        _safe_number(report.total_expenses)
        for report in reports
    )

    purchases = sum(
    _safe_number(report.total_purchases)
    for report in reports
)

    bills = sum(
        report.total_bills or 0
        for report in reports
    )

    deliveries = sum(
        report.deliveries or 0
        for report in reports
    )

    return {

        "sales": round(sales, 2),

        "expenses": round(expenses, 2),

        "purchases": round(purchases, 2),

        "bills": bills,

        "deliveries": deliveries,

    }

@router.get("/export/excel")
def export_excel(
    period: str = "today",
    store_id: str = "all",
    from_date: str | None = None,
    to_date: str | None = None,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):

    require_role(
        ["owner", "store_manager"],
        current_user["role"],
    )

    if current_user["role"] == "store_manager":
        store_id = str(current_user["store_id"])

    reports = _get_reports(
    db=db,
    period=period,
    store_id=store_id,
    from_date=from_date,
    to_date=to_date,
    submitted_only=current_user["role"] == "owner",
)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Analytics"

    headers = [
        "Store",
        "Date",
        "Bills",
        "Sales",
        "Expenses",
        "Purchases",
    ]

    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column)
        cell.value = header
        cell.font = Font(bold=True)

    row = 2

    for report in reports:

        store = (
            db.query(Store)
            .filter(Store.id == report.store_id)
            .first()
        )

        purchases = report.total_purchases

        sales = (
            _safe_number(report.cash_sales)
            + _safe_number(report.upi_sales)
            + _safe_number(report.card_sales)
            + _safe_number(report.udhaar_sales)
        )

        sheet.cell(row=row, column=1).value = (
            store.name if store else ""
        )

        sheet.cell(row=row, column=2).value = str(
            report.report_date
        )

        sheet.cell(row=row, column=3).value = (
            report.total_bills
        )

        sheet.cell(row=row, column=4).value = sales

        sheet.cell(row=row, column=5).value = (
            report.total_expenses
        )

        sheet.cell(row=row, column=6).value = purchases

        row += 1

    stream = BytesIO()

    workbook.save(stream)

    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
            "attachment; filename=analytics.xlsx"
        },
    )


@router.get("/export/pdf")
def export_pdf(
    period: str = "today",
    store_id: str = "all",
    from_date: str | None = None,
to_date: str | None = None,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):

    require_role(
        ["owner", "store_manager"],
        current_user["role"],
    )

    if current_user["role"] == "store_manager":
        store_id = str(current_user["store_id"])

    reports = _get_reports(
    db=db,
    period=period,
    store_id=store_id,
    from_date=from_date,
    to_date=to_date,
    submitted_only=current_user["role"] == "owner",
)
    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
    )

    table_data = [[
        "Store",
        "Date",
        "Bills",
        "Sales",
        "Expenses",
        "Purchases",
    ]]

    for report in reports:

        store = (
            db.query(Store)
            .filter(Store.id == report.store_id)
            .first()
        )

        purchases = (
            db.query(
                func.sum(
                    Purchase.purchase_amount
                )
            )
            .filter(
                Purchase.store_id == report.store_id
            )
            .scalar()
            or 0
        )

        sales = (
            _safe_number(report.cash_sales)
            + _safe_number(report.upi_sales)
            + _safe_number(report.card_sales)
            + _safe_number(report.udhaar_sales)
        )

        table_data.append([
            store.name if store else "",
            str(report.report_date),
            report.total_bills,
            round(sales, 2),
            round(
                _safe_number(
                    report.total_expenses
                ),
                2,
            ),
            round(purchases, 2),
        ])

    table = Table(table_data)

    table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#2563EB"),
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white,
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey,
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold",
            ),
            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, 0),
                10,
            ),
        ])
    )

    document.build([table])

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
            "attachment; filename=analytics.pdf"
        },
    )
